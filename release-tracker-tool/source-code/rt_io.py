"""
Release Tracker — tabular I/O (CSV + Excel) for Import, Export and Bulk Update.

Reading and writing both formats is centralised here so the routes stay thin.
Excel uses ``openpyxl`` (pure-Python, added to requirements); CSV uses the stdlib
``csv`` module. Reads are defensive (BOM-tolerant, blank-row-skipping); writes
preserve real date typing and apply a light, consistent style so exports look
professional in Excel.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Columns emitted by Export, in order: (internal key, friendly header).
EXPORT_FIELDS = [
    ("s_no", "S.No"),
    ("enhancement_id", "Enhancement/Case ID"),
    ("release_subject", "Release Mail Subject"),
    ("category", "Category"),
    ("other_category", "Other Category"),
    ("sent_by", "Sent By"),
    ("batch_number", "Batch Number"),
    ("crm_delivery_date", "CRM Delivery Date"),
    ("sit_date", "SIT Execution Date"),
    ("uat_date", "UAT Execution Date"),
    ("preprod_date", "PreProd Execution Date"),
    ("prod_live_date", "Prod Live Date"),
    ("created_by", "Created By"),
    ("created_date", "Created Date"),
    ("updated_by", "Updated By"),
    ("updated_date", "Updated Date"),
]

_DATE_KEYS = {"crm_delivery_date", "sit_date", "uat_date", "preprod_date", "prod_live_date"}
_HEADER_FILL = "4F46E5"   # brand-600


class ImportFormatError(ValueError):
    """Raised when an uploaded file cannot be read as CSV/Excel."""


# ----------------------------------------------------------------------
# Reading
# ----------------------------------------------------------------------
def _read_csv(data: bytes) -> list[list]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


def _read_xlsx(data: bytes) -> list[list]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ImportFormatError(f"Could not read the Excel file: {exc}") from exc
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def read_table(filename: str, data: bytes) -> list[list]:
    """Return all rows (including the header row) as a list of cell lists."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        rows = _read_csv(data)
    elif name.endswith((".xlsx", ".xlsm")):
        rows = _read_xlsx(data)
    else:
        raise ImportFormatError("Unsupported file type. Upload a .csv or .xlsx file.")
    # Drop fully-empty rows.
    return [r for r in rows if any(str(c).strip() for c in r if c is not None)]


def read_records(filename: str, data: bytes) -> tuple[list[dict], dict]:
    """Read an upload into ``(raw_rows, header_map)``.

    ``raw_rows`` are dicts keyed by internal column name (only mapped columns are
    present); ``header_map`` is the resolved ``{column: index}``. The first
    non-empty row is treated as the header.
    """
    from rt_service import map_headers

    rows = read_table(filename, data)
    if not rows:
        return [], {}
    header_map = map_headers(rows[0])
    raw = []
    for cells in rows[1:]:
        rec = {}
        for col, idx in header_map.items():
            rec[col] = cells[idx] if idx < len(cells) else None
        raw.append(rec)
    return raw, header_map


# ----------------------------------------------------------------------
# Writing — Export
# ----------------------------------------------------------------------
def _to_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, _dt.date):
        return v
    try:
        return _dt.datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return v


def export_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([label for _k, label in EXPORT_FIELDS])
    for r in rows:
        w.writerow([r.get(k, "") if r.get(k) is not None else "" for k, _label in EXPORT_FIELDS])
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx(rows: list[dict], *, sheet_title: str = "Release Tracker") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Release Tracker"

    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    center = Alignment(horizontal="center", vertical="center")

    labels = [label for _k, label in EXPORT_FIELDS]
    ws.append(labels)
    for col_idx, _label in enumerate(labels, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = fill
        c.alignment = center

    for r in rows:
        out_row = []
        for k, _label in EXPORT_FIELDS:
            v = r.get(k)
            out_row.append(_to_date(v) if k in _DATE_KEYS else v)
        ws.append(out_row)

    # Apply a date number format + reasonable widths.
    for col_idx, (k, label) in enumerate(EXPORT_FIELDS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(12, min(40, len(label) + 4))
        if k in _DATE_KEYS:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"

    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_bytes(rows: list[dict], fmt: str, *, sheet_title: str = "Release Tracker") -> tuple[bytes, str, str]:
    """Return ``(payload, mimetype, extension)`` for the requested format."""
    if fmt == "csv":
        return (export_csv(rows), "text/csv", "csv")
    return (export_xlsx(rows, sheet_title=sheet_title),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx")


def error_report_csv(errors: list[dict]) -> bytes:
    """Build a downloadable error report from ``[{row, batch, error}, ...]``."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Row", "Batch Number", "Error"])
    for e in errors:
        w.writerow([e.get("row", ""), e.get("batch", ""), e.get("error", "")])
    return buf.getvalue().encode("utf-8-sig")
